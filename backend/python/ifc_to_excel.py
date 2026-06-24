"""
IFC → GMAP Material Passport  Excel + CSV Converter   v 3.1
============================================================
Standalone script — no FastAPI, no web server required.

Run:
    pip install ifcopenshell openpyxl trimesh numpy
    # optional, enables the OpenCASCADE geometry fallback:
    pip install pythonocc-core

    python ifc_to_passport_v3_1.py               # opens file-picker dialog
    python ifc_to_passport_v3_1.py ./project.ifc  # direct path

Changelog vs v3.0
    ✓ ARCHITECTURAL FIX: One row per material constituent — not one row per
      element. "Brick | EPS | Plaster" crammed into a single cell is gone.
      An element with N layers/constituents now produces N passport rows,
      each with its own material name, thickness, proportioned volume and
      mass. Layer Build-up retains the full assembly string on every row
      for assembly context.
    ✓ OCC import rewritten with importlib. The direct
      `from OCC.Core.GProp import GProp_GProps` raised a Pylance
      "could not be resolved" false-positive; importlib-based dynamic
      loading fixes this without changing runtime behaviour.
    ✓ `get_all_psets` helper: merges instance-level AND type-level psets
      so engineering/lifecycle properties defined on IfcWallType (the
      common practice) are no longer silently missed.
    ✓ Fire rating now searched across 5 key name variants in every pset,
      not just `Pset_WallCommon.FireRating`.
    ✓ Thermal transmittance, acoustic rating likewise expanded with alias
      lists (ThermalTransmittance / U-Value / ThermalResistance / R-Value
      etc.).
    ✓ Lifecycle/circularity fields look in 8 additional psets:
      Pset_ServiceLife, Pset_Warranty, Pset_Asset, Pset_Condition,
      custom "Madaster_*" variants, etc.
    ✓ Density resolved per-constituent: IFC HasProperties first, then
      DENSITY_LOOKUP — so composite element mass is correctly proportioned
      row-by-row instead of using a single averaged value.
    ✓ GMAP Id is constituent-aware: single-material elements produce the
      same ID as v3.0; multi-constituent elements get a unique ID per
      constituent row (deterministic; re-exports are stable).
    ✓ Summary sheet "Total Elements (IFC)" now shows unique element count
      separately from total passport rows.
"""

from __future__ import annotations

import csv
import datetime
import importlib
import os
import re
import sys
import threading
import traceback
import uuid
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ── UTF-8 console output (Windows / child_process safe) ──────────────────
try:
    # line_buffering=True: every \n flushes automatically — critical for pipes
    sys.stdout.reconfigure(encoding='utf-8', line_buffering=True)
    sys.stderr.reconfigure(encoding='utf-8', line_buffering=True)
except Exception:
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

# ── dependency guard (hard requirements) ───────────────────────────────────
def _require(import_name: str, pip_name: str = "") -> None:
    try:
        __import__(import_name)
    except ImportError:
        pkg = pip_name or import_name
        print(f"\n[ERROR] Missing package - run:  pip install {pkg}\n", file=sys.stderr)
        sys.exit(1)

_require("ifcopenshell")
_require("openpyxl")
_require("numpy")
_require("trimesh")

import ifcopenshell
import ifcopenshell.util.element        as ifc_elem
import ifcopenshell.util.unit           as ifc_unit
import ifcopenshell.util.classification as ifc_cls
import ifcopenshell.geom                as ifc_geom
import numpy   as np
import trimesh as tr
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── optional dependency: pythonocc-core (OpenCASCADE fallback) ─────────────
# Loaded via importlib to avoid Pylance "could not be resolved" false-positive.
# Runtime behaviour is identical — missing package is handled gracefully.
_OCC_AVAILABLE       = False
_GProp_GProps        = None            # type: ignore[assignment]
_brepgprop_VolProps  = None            # type: ignore[assignment]

try:
    _gprop_mod = importlib.import_module("OCC.Core.GProp")
    _GProp_GProps = _gprop_mod.GProp_GProps                         # type: ignore[attr-defined]
    _brep_mod = importlib.import_module("OCC.Core.BRepGProp")
    _brepgprop_VolProps = getattr(_brep_mod, "brepgprop_VolumeProperties", None)
    if _brepgprop_VolProps is None:
        _inner = getattr(_brep_mod, "brepgprop", None)
        if _inner:
            _brepgprop_VolProps = getattr(_inner, "VolumeProperties", None)
    if _GProp_GProps is not None and _brepgprop_VolProps is not None:
        _OCC_AVAILABLE = True
except Exception:
    pass


# =============================================================================
# 1. COLUMN SCHEMA  (50 columns, 9 groups) — UNCHANGED from v3.0
# =============================================================================
COLUMN_GROUPS: Dict[str, Tuple[str, str, List[str]]] = {
    "Identification":       ("1F3864", "BDD7EE", [
        "GMAP Id", "IFC GUID", "GTIN", "Article Number", "External DB Id",
    ]),
    "Element":              ("375623", "E2EFDA", [
        "IFC Class", "Type Name", "Element Name", "Description", "Floor / Storey", "Status",
    ]),
    "Material Composition": ("7B3206", "FCE4D6", [
        "Material / Product", "Material Type", "Layer Build-up", "Material Category",
    ]),
    "Quantities":           ("4A235A", "EAD1DC", [
        "Volume (m³)", "Geom Volume (m³)", "Vol Check",
        "Area (m²)", "Length (m)", "Height (m)", "Width (m)", "Thickness (m)", "Diameter (m)",
    ]),
    "Mass & Carbon":        ("7B6000", "FFF2CC", [
        "Mass (kg)", "Density (kg/m³)",
        "Embodied Carbon A1-A3 (kg CO\u2082e)", "GWP / kg (kg CO\u2082e/kg)",
    ]),
    "Engineering":          ("1A5276", "D6EAF8", [
        "Is External", "Load Bearing", "Fire Rating",
        "Thermal Transmittance (W/m\u00b2K)", "Acoustic Rating (dB)",
    ]),
    "Classification":       ("1D6A54", "D9EAD3", [
        "Classification System", "Classification Code", "Classification Name",
    ]),
    "Circularity & EOL":    ("78281F", "F4CCCC", [
        "% Reused", "% Available for Reuse", "Assumed Construction Waste", "Waste Codes",
        "Detachability \u2013 Connection", "Detachability \u2013 Accessibility",
        "Detachability \u2013 Intersection", "Detachability \u2013 Product Edge",
    ]),
    "Lifecycle & Asset":    ("4D4D4D", "F2F2F2", [
        "Lifespan (Years)", "Installation Date", "Serial Number",
        "Technical Condition", "Aesthetic Condition", "Comment",
    ]),
}

ALL_COLS: List[str] = [c for _, _, cols in COLUMN_GROUPS.values() for c in cols]

COL_WIDTHS: Dict[str, float] = {
    "GMAP Id":                            24,
    "IFC GUID":                           26,
    "GTIN":                               16,
    "Article Number":                     18,
    "External DB Id":                     20,
    "IFC Class":                          26,
    "Type Name":                          30,
    "Element Name":                       24,
    "Description":                        48,
    "Floor / Storey":                     18,
    "Status":                             13,
    "Material / Product":                 36,
    "Material Type":                      28,
    "Layer Build-up":                     64,
    "Material Category":                  20,
    "Volume (m\u00b3)":                   14,
    "Geom Volume (m\u00b3)":              16,
    "Vol Check":                          14,
    "Area (m\u00b2)":                     14,
    "Length (m)":                         13,
    "Height (m)":                         13,
    "Width (m)":                          13,
    "Thickness (m)":                      14,
    "Diameter (m)":                       14,
    "Mass (kg)":                          14,
    "Density (kg/m\u00b3)":              16,
    "Embodied Carbon A1-A3 (kg CO\u2082e)": 30,
    "GWP / kg (kg CO\u2082e/kg)":         22,
    "Is External":                        14,
    "Load Bearing":                       14,
    "Fire Rating":                        14,
    "Thermal Transmittance (W/m\u00b2K)": 28,
    "Acoustic Rating (dB)":              22,
    "Classification System":              24,
    "Classification Code":               22,
    "Classification Name":               28,
    "% Reused":                           12,
    "% Available for Reuse":             22,
    "Assumed Construction Waste":        28,
    "Waste Codes":                        16,
    "Detachability \u2013 Connection":   26,
    "Detachability \u2013 Accessibility": 28,
    "Detachability \u2013 Intersection": 26,
    "Detachability \u2013 Product Edge": 26,
    "Lifespan (Years)":                  16,
    "Installation Date":                 18,
    "Serial Number":                     18,
    "Technical Condition":               22,
    "Aesthetic Condition":               22,
    "Comment":                           38,
}

NUM_FMT: Dict[str, str] = {
    "Volume (m\u00b3)":                  "0.00000",
    "Geom Volume (m\u00b3)":             "0.00000",
    "Area (m\u00b2)":                    "0.00000",
    "Length (m)":                        "0.00000",
    "Height (m)":                        "0.00000",
    "Width (m)":                         "0.00000",
    "Thickness (m)":                     "0.00000",
    "Diameter (m)":                      "0.00000",
    "Mass (kg)":                         "#,##0.00",
    "Density (kg/m\u00b3)":             "#,##0.00",
    "Embodied Carbon A1-A3 (kg CO\u2082e)": "#,##0.000",
    "GWP / kg (kg CO\u2082e/kg)":        "0.0000",
    "Thermal Transmittance (W/m\u00b2K)": "0.000",
    "Acoustic Rating (dB)":             "0.0",
    "Lifespan (Years)":                  "0",
}

NUMERIC_COLS = set(NUM_FMT.keys())
WRAP_COLS = {"Description", "Layer Build-up", "Material / Product", "Comment"}
EXCEL_CELL_CHAR_LIMIT = 32000

TARGET_IFC_TYPES: List[str] = [
    "IfcWall", "IfcWallStandardCase",
    "IfcSlab",
    "IfcBeam",
    "IfcColumn",
    "IfcFooting",
    "IfcPile",
    "IfcRoof",
    "IfcStair", "IfcStairFlight",
    "IfcRamp", "IfcRampFlight",
    "IfcDoor",
    "IfcWindow",
    "IfcPlate",
    "IfcMember",
    "IfcCovering",
    "IfcRailing",
    "IfcFurnishingElement",
    "IfcBuildingElementProxy",
    "IfcFlowSegment", "IfcFlowTerminal", "IfcFlowFitting",
]

# Expanded — includes type-level psets + sub-element psets
ENG_PSETS: List[str] = [
    "Pset_WallCommon", "Pset_SlabCommon", "Pset_BeamCommon",
    "Pset_ColumnCommon", "Pset_DoorCommon", "Pset_WindowCommon",
    "Pset_RoofCommon", "Pset_StairCommon", "Pset_RampCommon",
    "Pset_CoveringCommon", "Pset_MemberCommon", "Pset_PlateCommon",
    "Pset_BuildingElementProxyCommon", "Pset_FootingCommon",
    "Pset_PileCommon", "Pset_RailingCommon",
    "Pset_ConcreteElementGeneral", "Pset_StructuralSurfaceMemberCommon",
    "Pset_StructuralCurveMemberCommon",
]

# Property key aliases (first match wins across all psets)
_FIRE_KEYS  = ["FireRating", "FireResistanceRating", "FireResistanceClass",
               "FireProtectionClass", "FireEscapeRating", "FireResistance"]
_THERM_KEYS = ["ThermalTransmittance", "ThermalTransmittance(U)", "U-Value",
               "UValue", "ThermalResistance", "R-Value", "RValue"]
_ACOU_KEYS  = ["AcousticRating", "SoundTransmissionClass", "STC",
               "AcousticPerformance", "SoundReduction", "Rw"]

DETACH_PSETS: Tuple[str, ...] = (
    "Pset_Disassembly", "Pset_Madaster", "Madaster_Disassembly",
    "Detachability", "Pset_Recyclability", "Pset_WasteInformation",
    "Pset_DisassemblyInformation",
)

# Additional psets for lifecycle / circularity / condition
LIFECYCLE_PSETS: Tuple[str, ...] = (
    "Pset_LifeCycleInformation", "Pset_ServiceLife", "Pset_Warranty",
    "Pset_Asset", "Pset_Condition", "Pset_Sustainability",
    "Madaster_LifeCycle", "Pset_CircularityInformation",
)

_CLASS_LABELS: Dict[str, str] = {
    "IfcWall": "Wall", "IfcWallStandardCase": "Wall", "IfcSlab": "Slab",
    "IfcBeam": "Beam", "IfcColumn": "Column", "IfcFooting": "Footing",
    "IfcPile": "Pile", "IfcRoof": "Roof", "IfcStair": "Stair",
    "IfcStairFlight": "Stair Flight", "IfcRamp": "Ramp", "IfcRampFlight": "Ramp Flight",
    "IfcDoor": "Door", "IfcWindow": "Window", "IfcPlate": "Plate",
    "IfcMember": "Member", "IfcCovering": "Covering", "IfcRailing": "Railing",
    "IfcFurnishingElement": "Furnishing", "IfcBuildingElementProxy": "Building Element",
    "IfcFlowSegment": "Flow Segment", "IfcFlowTerminal": "Flow Terminal",
    "IfcFlowFitting": "Flow Fitting",
}


# =============================================================================
# 2. MATERIAL LOOKUPS
# =============================================================================
DENSITY_LOOKUP: Dict[str, float] = {
    "rcc m25": 2400, "rcc m30": 2500, "rcc": 2500,
    "reinforced concrete": 2500, "pcc": 2300, "concrete": 2400,
    "fired clay brick": 1800, "fly ash brick": 1700, "brick": 1800,
    "aac block": 700, "autoclaved aerated": 700,
    "stone masonry": 2200, "rubble masonry": 2100,
    "structural steel s355": 7850, "structural steel": 7850,
    "mild steel": 7850, "stainless steel": 7900,
    "steel": 7850, "aluminium": 2700, "aluminum": 2700,
    "copper": 8960, "iron": 7874,
    "teak": 650, "bamboo": 400, "plywood": 600, "timber": 600, "wood": 600,
    "float glass": 2500, "glass": 2500,
    "cement plaster": 1800, "cement mortar": 2000,
    "plaster": 1800, "mortar": 2000,
    "granite": 2700, "marble": 2700, "sandstone": 2200, "limestone": 2500,
    "glass wool": 30, "rock wool": 80, "eps": 25, "xps": 35,
    "gypsum board": 900, "gypsum": 1200,
    "paint": 1200, "waterproofing": 1100,
    "rebar": 7850, "reinforcement": 7850,
    "ceramic": 1900, "tile": 1900, "porcelain": 2300,
    "pvc": 1400, "upvc": 1400, "hdpe": 960, "polypropylene": 910,
}

CATEGORY_LOOKUP: Dict[str, str] = {
    "rcc": "Structural", "reinforced concrete": "Structural",
    "concrete": "Structural", "pcc": "Structural",
    "structural steel": "Structural", "steel": "Structural",
    "iron": "Structural", "aluminium": "Structural", "aluminum": "Structural",
    "rebar": "Structural", "reinforcement": "Structural",
    "brick": "Masonry", "block": "Masonry", "stone masonry": "Masonry",
    "glass": "Facade", "float glass": "Facade",
    "wood": "Timber", "timber": "Timber", "teak": "Timber",
    "plywood": "Timber", "bamboo": "Timber",
    "plaster": "Finishing", "gypsum board": "Finishing",
    "paint": "Finishing", "tile": "Finishing",
    "ceramic": "Finishing", "porcelain": "Finishing",
    "glass wool": "Insulation", "rock wool": "Insulation",
    "eps": "Insulation", "xps": "Insulation",
    "granite": "Stone", "marble": "Stone",
    "sandstone": "Stone", "limestone": "Stone",
    "copper": "MEP", "pvc": "MEP", "upvc": "MEP", "hdpe": "MEP",
}


def _mat_lookup(name: str, table: Dict[str, Any]) -> Optional[Any]:
    """Case-insensitive substring lookup; longest matching key wins."""
    low = (name or "").lower()
    best_key, best_val = "", None
    for k, v in table.items():
        if k in low and len(k) > len(best_key):
            best_key, best_val = k, v
    return best_val


# =============================================================================
# 3. GMAP ID GENERATION  — constituent-aware
# =============================================================================
_GMAP_NAMESPACE = uuid.uuid5(uuid.NAMESPACE_URL, "https://gmap.invalid/material-passport")


def generate_gmap_id(project_name: str, ifc_guid: str,
                     constituent_index: int = 0, total_constituents: int = 1) -> str:
    """
    Deterministic GMAP Id, distinct from the IFC GUID.
    For single-material elements (total_constituents == 1) the ID is
    identical to v3.0 output.  Multi-constituent elements get a unique
    ID per row so each passport entry is independently addressable.
    """
    try:
        key = (f"{project_name}|{ifc_guid}"
               if total_constituents == 1
               else f"{project_name}|{ifc_guid}|{constituent_index}")
        gid = uuid.uuid5(_GMAP_NAMESPACE, key)
        return f"GMAP-{str(gid).upper().replace('-', '')[:20]}"
    except Exception:
        return f"GMAP-{abs(hash(ifc_guid + str(constituent_index))) & 0xFFFFFFFFFFFF:012X}"


# =============================================================================
# 4. UNIT SCALE
# =============================================================================
def get_unit_scale(model) -> float:
    try:
        s = ifc_unit.calculate_unit_scale(model)
        return float(s) if s else 1.0
    except Exception:
        return 1.0


# =============================================================================
# 5. STOREY NORMALISATION
# =============================================================================
_GENERIC_LEVEL_RE = re.compile(
    r'^\s*(level|storey|story|floor|lvl|fl|l)\.?\s*[-_]?\s*0*\d+\s*$', re.IGNORECASE)

_ORDINAL_FLOOR_NAMES = [
    "Ground Floor", "First Floor", "Second Floor", "Third Floor", "Fourth Floor",
    "Fifth Floor", "Sixth Floor", "Seventh Floor", "Eighth Floor", "Ninth Floor",
    "Tenth Floor", "Eleventh Floor", "Twelfth Floor",
]


def _ordinal_floor_name(n: int) -> str:
    if 0 <= n < len(_ORDINAL_FLOOR_NAMES):
        return _ORDINAL_FLOOR_NAMES[n]
    suffix = "th"
    if n % 10 == 1 and n % 100 != 11: suffix = "st"
    elif n % 10 == 2 and n % 100 != 12: suffix = "nd"
    elif n % 10 == 3 and n % 100 != 13: suffix = "rd"
    return f"{n + 1}{suffix} Floor"


def _looks_generic(name: str) -> bool:
    return bool(_GENERIC_LEVEL_RE.match(name or "")) or not (name or "").strip()


def _tidy_name(name: str) -> str:
    name = name.strip()
    return name.title() if (name.isupper() or name.islower()) else name


def build_storey_labels(model) -> Dict[str, str]:
    try:
        storeys = model.by_type("IfcBuildingStorey")
    except Exception:
        storeys = []
    if not storeys:
        return {}

    def _elev(s) -> float:
        try:
            e = getattr(s, "Elevation", None)
            return float(e) if e is not None else 0.0
        except Exception:
            return 0.0

    try:
        ordered = sorted(storeys, key=_elev)
    except Exception:
        ordered = list(storeys)

    above = [s for s in ordered if _elev(s) >= -0.5]
    below = list(reversed([s for s in ordered if _elev(s) < -0.5]))

    labels: Dict[str, str] = {}
    for i, s in enumerate(above):
        raw = (getattr(s, "Name", None) or getattr(s, "LongName", None) or "").strip()
        labels[s.GlobalId] = _ordinal_floor_name(i) if _looks_generic(raw) else _tidy_name(raw)
    for i, s in enumerate(below):
        raw = (getattr(s, "Name", None) or getattr(s, "LongName", None) or "").strip()
        labels[s.GlobalId] = f"Basement {i + 1}" if _looks_generic(raw) else _tidy_name(raw)
    return labels


def build_storey_map(model) -> Dict[str, str]:
    storey_labels = build_storey_labels(model)
    guid_to_storey: Dict[str, str] = {}

    def _walk(entity, current_label: str = "") -> None:
        label = current_label
        if entity.is_a("IfcBuildingStorey"):
            label = storey_labels.get(entity.GlobalId,
                                       getattr(entity, "Name", "") or current_label)
        for rel in getattr(entity, "ContainsElements", []) or []:
            for el in rel.RelatedElements:
                guid_to_storey[el.GlobalId] = label
        for rel in getattr(entity, "IsDecomposedBy", []) or []:
            for child in rel.RelatedObjects:
                _walk(child, label)

    for proj in model.by_type("IfcProject"):
        try:
            _walk(proj)
        except Exception as e:
            print(f"  [WARNING] Spatial tree walk error: {e}")

    for rel in model.by_type("IfcRelContainedInSpatialStructure"):
        try:
            ss = rel.RelatingStructure
            if ss.is_a("IfcBuildingStorey"):
                label = storey_labels.get(ss.GlobalId, getattr(ss, "Name", "") or "")
                for el in rel.RelatedElements:
                    if el.GlobalId not in guid_to_storey:
                        guid_to_storey[el.GlobalId] = label
        except Exception:
            continue
    return guid_to_storey


def get_building_name(model) -> str:
    try:
        buildings = model.by_type("IfcBuilding")
        return (getattr(buildings[0], "Name", None) or "") if buildings else ""
    except Exception:
        return ""


# =============================================================================
# 6. PSET HELPER — merges instance + type psets (NEW in v3.1)
# =============================================================================
def get_all_psets(element) -> dict:
    """
    Returns merged pset dict from the element instance AND its IfcTypeObject.
    Type-level psets (e.g. on IfcWallType) are very common for fire rating,
    thermal transmittance, acoustic ratings, and classification.
    Instance psets take precedence if the same pset exists at both levels.
    """
    psets: dict = {}
    # Type first so instance-level values win when keys collide
    try:
        type_obj = ifc_elem.get_type(element)
        if type_obj is not None:
            psets.update(ifc_elem.get_psets(type_obj))
    except Exception:
        pass
    try:
        psets.update(ifc_elem.get_psets(element))
    except Exception:
        pass
    return psets


def _pset_pick(psets: dict, *keys: str) -> Optional[Any]:
    """Search ALL psets for any of the given property keys; first hit wins."""
    _null = {None, "", "N/A", "NOTDEFINED", "UNSET", "NOTKNOWN"}
    for k in keys:
        for pset_vals in psets.values():
            if isinstance(pset_vals, dict):
                v = pset_vals.get(k)
                if v not in _null:
                    return v
    return None


# =============================================================================
# 7. BATCH GEOMETRY
# =============================================================================
def extract_geometry(model, skip_geom: bool = False) -> Dict[str, dict]:
    """
    Extracts trimesh geometry for all IFC elements and returns a guid -> dict map.

    skip_geom=True returns {} immediately and is the recommended mode when called
    from an automated backend because the C++ batch iterator can deadlock on
    certain IFC files.  All material/mass data still comes from QTO psets.

    When geometry IS requested the function tries two strategies:
      1. Threaded batch ifc_geom.iterator() with a 90-second timeout.
         A daemon thread runs the C++ iterator; if it hangs past the timeout
         the thread is abandoned (it dies when the process exits) and we fall
         through to strategy 2.
      2. Per-element ifc_geom.create_shape() — slower but always terminates.
         Individual element failures are silently skipped.
    """
    if skip_geom:
        print("  [INFO] Geometry extraction disabled (--no-geom)")
        return {}

    geom_map: Dict[str, dict] = {}
    settings = ifc_geom.settings()
    settings.set(settings.USE_WORLD_COORDS, True)

    def _store(guid: str, verts_flat, faces_flat) -> None:
        try:
            verts = np.array(verts_flat, dtype=np.float64).reshape(-1, 3)
            faces = np.array(faces_flat, dtype=np.int32).reshape(-1, 3)
            if len(verts) < 4 or len(faces) < 1:
                return
            mesh = tr.Trimesh(vertices=verts, faces=faces, process=False)
            watertight = bool(mesh.is_watertight)
            vol = abs(float(mesh.volume)) if watertight else _divergence_vol(verts, faces)
            if vol < 1e-12:
                vol = _divergence_vol(verts, faces)
            bbox = mesh.bounding_box.bounds
            geom_map[guid] = {
                "geom_volume":   round(max(vol, 0.0), 6),
                "surface_area":  round(float(mesh.area), 6),
                "bbox_dims":     bbox[1] - bbox[0],
                "is_watertight": watertight,
            }
        except Exception as e:
            print(f"  [WARNING] Geometry error for {guid}: {e}")

    # ── Strategy 1: batch iterator with threading timeout ─────────────────
    # Cap threads at 4 to lower the chance of C++ deadlocks on Windows.
    n_threads     = min(os.cpu_count() or 4, 4)
    BATCH_TIMEOUT = 90  # seconds
    print(f"  Batch geometry: {n_threads} threads, {BATCH_TIMEOUT}s timeout...")

    batch_done:  threading.Event = threading.Event()
    batch_error: List[Exception] = []

    def _run_batch() -> None:
        try:
            try:
                it = ifc_geom.iterator(settings, model, num_threads=n_threads)
            except TypeError:
                it = ifc_geom.iterator(settings, model, n_threads)
            if it.initialize():
                while True:
                    s = it.get()
                    _store(s.guid, s.geometry.verts, s.geometry.faces)
                    if not it.next():
                        break
        except Exception as exc:
            batch_error.append(exc)
        finally:
            batch_done.set()

    batch_thread = threading.Thread(target=_run_batch, daemon=True)
    batch_thread.start()

    if batch_done.wait(timeout=BATCH_TIMEOUT):
        if not batch_error:
            print(f"  [OK] Geometry: {len(geom_map)} meshes ({n_threads} threads)")
            return geom_map
        print(f"  [WARNING] Batch geometry failed: {batch_error[0]} - falling back to per-element")
    else:
        print(f"  [WARNING] Batch geometry timed out after {BATCH_TIMEOUT}s - falling back to per-element")

    # Clear any partial results from the (still-running daemon) batch thread
    geom_map.clear()

    # ── Strategy 2: per-element fallback ──────────────────────────────────
    print("  Per-element geometry fallback (slower, always terminates)...")
    all_els = list(model.by_type("IfcElement"))
    total_g = len(all_els)
    for gi, el in enumerate(all_els):
        guid = getattr(el, "GlobalId", None)
        if not guid:
            continue
        if gi % 50 == 0 or gi == total_g - 1:
            print(f"    Geom {gi + 1}/{total_g}...")
        try:
            s = ifc_geom.create_shape(settings, el)
            _store(guid, s.geometry.verts, s.geometry.faces)
        except Exception:
            pass  # geometry is optional — silently skip failed elements

    print(f"  [OK] Geometry (per-element): {len(geom_map)}/{total_g} meshes")
    return geom_map


def _divergence_vol(verts: np.ndarray, faces: np.ndarray) -> float:
    vol = 0.0
    for f in faces:
        v0, v1, v2 = verts[f[0]], verts[f[1]], verts[f[2]]
        vol += float(np.dot(v0, np.cross(v1, v2))) / 6.0
    return abs(vol)


def occ_volume_fallback(element) -> Optional[float]:
    if not _OCC_AVAILABLE:
        return None
    try:
        occ_settings = ifc_geom.settings()
        occ_settings.set(occ_settings.USE_PYTHON_OPENCASCADE, True)
        shape = ifc_geom.create_shape(occ_settings, element)
        topods_shape = shape.geometry
        props = _GProp_GProps()
        _brepgprop_VolProps(topods_shape, props)
        vol = abs(props.Mass())
        return round(vol, 6) if vol > 1e-9 else None
    except AttributeError:
        return None
    except Exception as e:
        print(f"  [WARNING] OCC fallback failed for {getattr(element, 'GlobalId', '?')}: {e}")
        return None


# =============================================================================
# 8. MATERIAL CONSTITUENTS  (CORE FIX: replaces get_material_info)
# =============================================================================
# Each constituent dict:
#   name          – single resolved material name (NOT pipe-joined)
#   ifc_material  – IfcMaterial object for IFC-native density lookup
#   mat_type      – e.g. "IfcMaterialLayer", "IfcMaterialConstituent"
#   layer_info    – positional label: "Layer 2/3 · 200mm"
#   full_assembly – full assembly string for Layer Build-up column context
#   fraction      – 0-1 share of total element volume
#   thickness_m   – this layer's thickness in metres (None if not applicable)
#   category      – material category string

def _cat(name: str) -> str:
    return _mat_lookup(name or "", CATEGORY_LOOKUP) or ""


def get_material_constituents(element, unit_scale: float) -> List[dict]:
    """
    Returns a list of constituent dicts — one entry per material layer,
    constituent, or profile. Guaranteed to return at least one entry.
    """
    try:
        return _get_constituents_inner(element, unit_scale)
    except Exception as e:
        print(f"  [WARNING] Constituent parse error {getattr(element, 'GlobalId', '?')}: {e}")
        return [_empty_constituent()]


def _empty_constituent() -> dict:
    return {"name": None, "ifc_material": None, "mat_type": "None",
            "layer_info": "", "full_assembly": "", "fraction": 1.0,
            "thickness_m": None, "category": ""}


def _make_c(name, ifc_mat, mat_type, layer_info, full_assembly, fraction, thickness_m) -> dict:
    return {
        "name": name or None,
        "ifc_material": ifc_mat,
        "mat_type": mat_type,
        "layer_info": layer_info,
        "full_assembly": full_assembly,
        "fraction": max(fraction, 0.0),
        "thickness_m": thickness_m,
        "category": _cat(name or ""),
    }


def _get_constituents_inner(element, unit_scale: float) -> List[dict]:
    try:
        mat = ifc_elem.get_material(element)
    except Exception:
        mat = None

    if mat is None:
        return [_empty_constituent()]

    # ── Unwrap usages ───────────────────────────────────────────────────────
    if mat.is_a("IfcMaterialLayerSetUsage"):
        mat = mat.ForLayerSet
    if mat.is_a("IfcMaterialProfileSetUsage") and getattr(mat, "ForProfileSet", None):
        mat = mat.ForProfileSet

    # ── Single material ─────────────────────────────────────────────────────
    if mat.is_a("IfcMaterial"):
        nm = mat.Name or ""
        mcat = getattr(mat, "Category", None) or _cat(nm)
        return [{
            "name": nm, "ifc_material": mat, "mat_type": "IfcMaterial",
            "layer_info": "", "full_assembly": nm,
            "fraction": 1.0, "thickness_m": None, "category": str(mcat),
        }]

    # ── Layer set ───────────────────────────────────────────────────────────
    if mat.is_a("IfcMaterialLayerSet"):
        layers = mat.MaterialLayers or []
        thicknesses = [max(float(getattr(l, "LayerThickness", None) or 0), 0.0) for l in layers]
        total_t = sum(thicknesses) or 1.0

        assembly_parts = []
        for l, t in zip(layers, thicknesses):
            m = getattr(l, "Material", None)
            nm = (m.Name if m else None) or "?"
            mm = int(round(t * unit_scale * 1000))
            assembly_parts.append(f"{nm} {mm}mm" if t > 0 else nm)
        full_assembly = " | ".join(assembly_parts) if assembly_parts else ""

        result, n = [], len(layers)
        for i, (l, t) in enumerate(zip(layers, thicknesses)):
            m = getattr(l, "Material", None)
            nm = (m.Name if m else None) or "?"
            mm = int(round(t * unit_scale * 1000))
            info = f"Layer {i+1}/{n}" + (f" · {mm}mm" if t > 0 else "")
            result.append(_make_c(nm, m, "IfcMaterialLayer", info, full_assembly,
                                  t / total_t, round(t * unit_scale, 6) if t > 0 else None))
        return result or [_empty_constituent()]

    # ── Constituent set ─────────────────────────────────────────────────────
    if mat.is_a("IfcMaterialConstituentSet"):
        parts = mat.MaterialConstituents or []
        fracs_raw = [getattr(c, "Fraction", None) for c in parts]
        has_fracs = any(f is not None for f in fracs_raw)
        n = len(parts) or 1
        total_f = (sum(float(f) for f in fracs_raw if f is not None) or 1.0) if has_fracs else n

        assembly_parts = []
        for c, f in zip(parts, fracs_raw):
            m = getattr(c, "Material", None)
            nm = (m.Name if m else None) or "?"
            pct = f" ({float(f)*100:.0f}%)" if f is not None else ""
            assembly_parts.append(f"{nm}{pct}")
        full_assembly = " | ".join(assembly_parts)

        result = []
        for i, (c, f_raw) in enumerate(zip(parts, fracs_raw)):
            m = getattr(c, "Material", None)
            nm = (m.Name if m else None) or "?"
            frac = (float(f_raw) / total_f if has_fracs and f_raw is not None else 1.0 / n)
            pct_str = f" · {float(f_raw)*100:.0f}%" if f_raw is not None else ""
            info = f"Constituent {i+1}/{n}{pct_str}"
            result.append(_make_c(nm, m, "IfcMaterialConstituent", info, full_assembly,
                                  frac, None))
        return result or [_empty_constituent()]

    # ── Profile set ─────────────────────────────────────────────────────────
    if mat.is_a("IfcMaterialProfileSet"):
        profs = mat.MaterialProfiles or []
        names_mats = []
        for p in profs:
            m = getattr(p, "Material", None)
            names_mats.append(((m.Name if m else None) or "?", m))
        full_assembly = " | ".join(nm for nm, _ in names_mats)
        n = len(profs) or 1
        result = []
        for i, (nm, m) in enumerate(names_mats):
            info = f"Profile {i+1}/{n}"
            result.append(_make_c(nm, m, "IfcMaterialProfile", info, full_assembly,
                                  1.0 / n, None))
        return result or [_empty_constituent()]

    # ── Legacy list ──────────────────────────────────────────────────────────
    if mat.is_a("IfcMaterialList"):
        mats_list = mat.Materials or []
        names_mats = [((m.Name or "?"), m) for m in mats_list]
        full_assembly = " | ".join(nm for nm, _ in names_mats)
        n = len(names_mats) or 1
        result = []
        for i, (nm, m) in enumerate(names_mats):
            info = f"Material {i+1}/{n}"
            result.append(_make_c(nm, m, "IfcMaterialList", info, full_assembly,
                                  1.0 / n, None))
        return result or [_empty_constituent()]

    return [_empty_constituent()]


# =============================================================================
# 9. PER-CONSTITUENT DENSITY  (replaces get_material_density)
# =============================================================================
def _density_from_ifc_material(ifc_mat) -> Optional[float]:
    """Read MassDensity from IFC material property sets."""
    if ifc_mat is None:
        return None
    for prop_set in getattr(ifc_mat, "HasProperties", []) or []:
        for prop in getattr(prop_set, "Properties", []) or []:
            key = (getattr(prop, "Name", "") or "").lower()
            if "density" in key or "massdensity" in key:
                nom = getattr(prop, "NominalValue", None)
                if nom is not None:
                    try:
                        v = float(nom.wrappedValue)
                        if v > 0:
                            return v
                    except Exception:
                        pass
    return None


def get_constituent_density(constituent: dict) -> Optional[float]:
    """
    Density (kg/m³) for a single constituent.
    IFC material property sets checked first; falls back to DENSITY_LOOKUP.
    """
    try:
        d = _density_from_ifc_material(constituent.get("ifc_material"))
        if d:
            return float(d)
        return _mat_lookup(constituent.get("name") or "", DENSITY_LOOKUP)
    except Exception:
        return None


# =============================================================================
# 10. TYPE NAME
# =============================================================================
def get_type_name(element) -> Optional[str]:
    try:
        type_obj = ifc_elem.get_type(element)
        if type_obj:
            return getattr(type_obj, "Name", None) or None
    except Exception:
        pass
    return None


# =============================================================================
# 11. CLASSIFICATION  (unchanged from v3.0)
# =============================================================================
def _walk_classification_root(ref) -> str:
    seen = set()
    current = ref
    while current is not None:
        try:
            cid = current.id()
        except Exception:
            break
        if cid in seen:
            break
        seen.add(cid)
        parent = getattr(current, "ReferencedSource", None)
        if parent is None:
            break
        if parent.is_a("IfcClassification"):
            return getattr(parent, "Name", "") or ""
        if parent.is_a("IfcClassificationReference"):
            current = parent
            continue
        break
    return ""


def get_classification(element) -> Tuple[str, str, str]:
    candidates = []

    def _collect(entity):
        try:
            for ref in ifc_cls.get_references(entity):
                candidates.append(ref)
        except Exception:
            pass
        for rel in getattr(entity, "HasAssociations", []) or []:
            try:
                if rel.is_a("IfcRelAssociatesClassification"):
                    ref = rel.RelatingClassification
                    if ref is not None and ref.is_a("IfcClassificationReference"):
                        candidates.append(ref)
            except Exception:
                continue

    try:
        _collect(element)
    except Exception:
        pass
    try:
        type_obj = ifc_elem.get_type(element)
        if type_obj is not None:
            _collect(type_obj)
    except Exception:
        pass

    if not candidates:
        return ("", "", "")

    seen_ids, uniq = set(), []
    for ref in candidates:
        try:
            rid = ref.id()
        except Exception:
            continue
        if rid not in seen_ids:
            seen_ids.add(rid)
            uniq.append(ref)

    if not uniq:
        return ("", "", "")

    ref = uniq[0]
    code   = getattr(ref, "Identification", None) or getattr(ref, "ItemReference", None) or ""
    label  = getattr(ref, "Name", "") or ""
    system = _walk_classification_root(ref)
    return (str(system), str(code), str(label))


# =============================================================================
# 12. QUANTITY EXTRACTION
# =============================================================================
def get_quantities(element, unit_scale: float) -> Dict[str, float]:
    result: Dict[str, float] = {}
    try:
        defs = getattr(element, "IsDefinedBy", []) or []
    except Exception:
        defs = []

    for rel in defs:
        try:
            if not rel.is_a("IfcRelDefinesByProperties"):
                continue
            defn = rel.RelatingPropertyDefinition
            if not defn.is_a("IfcElementQuantity"):
                continue
            for q in defn.Quantities:
                try:
                    name = (q.Name or "").lower()
                    if q.is_a("IfcQuantityLength"):
                        result[name] = q.LengthValue * unit_scale
                    elif q.is_a("IfcQuantityArea"):
                        result[name] = q.AreaValue * (unit_scale ** 2)
                    elif q.is_a("IfcQuantityVolume"):
                        result[name] = q.VolumeValue * (unit_scale ** 3)
                    elif q.is_a("IfcQuantityWeight"):
                        result[name] = float(q.WeightValue)
                    elif q.is_a("IfcQuantityCount"):
                        result[name] = float(q.CountValue)
                except Exception:
                    continue
        except Exception:
            continue
    return result


# =============================================================================
# 13. ENGINEERING PROPERTIES  (expanded: type psets + alias lists)
# =============================================================================
def get_engineering_props(psets: dict) -> dict:
    """
    Searches ALL psets (instance + type, merged by get_all_psets) using
    expanded alias lists.  First non-null value wins for each field.
    """
    merged: Dict[str, object] = {}
    for pname in ENG_PSETS:
        p = psets.get(pname)
        if isinstance(p, dict):
            merged.update(p)

    def _b(key: str) -> Optional[bool]:
        # Check merged standard psets first, then all psets
        v = merged.get(key)
        if v is None:
            v = _pset_pick(psets, key)
        return bool(v) if v is not None else None

    return {
        "Is External":    _b("IsExternal"),
        "Load Bearing":   _b("LoadBearing"),
        "Fire Rating":    _pset_pick(psets, *_FIRE_KEYS),
        "Thermal Transmittance (W/m\u00b2K)": _pset_pick(psets, *_THERM_KEYS),
        "Acoustic Rating (dB)":               _pset_pick(psets, *_ACOU_KEYS),
        "Status": (_pset_pick(psets, "Status", "ConstructionStatus",
                               "ObjectStatus", "Phase") or merged.get("Status")),
    }


def get_detachability(psets: dict) -> Tuple:
    for name in DETACH_PSETS:
        p = psets.get(name, {}) or {}
        if p:
            return (
                p.get("ConnectionType") or p.get("Connection") or p.get("JointType"),
                p.get("Accessibility"),
                p.get("Intersection"),
                p.get("ProductEdge") or p.get("ProductEdges"),
            )
    return (None, None, None, None)


# =============================================================================
# 14. DESCRIPTION BUILDER
# =============================================================================
def _humanize_class(ifc_class: str) -> str:
    if ifc_class in _CLASS_LABELS:
        return _CLASS_LABELS[ifc_class]
    base = ifc_class[3:] if ifc_class.startswith("Ifc") else ifc_class
    return re.sub(r'(?<!^)(?=[A-Z])', ' ', base).strip() or ifc_class


def build_description(ifc_class: str, name: str, desc: str, type_name: Optional[str],
                       mat_str: str, storey_label: Optional[str], eng: dict) -> str:
    try:
        label = _humanize_class(ifc_class)
        lead  = (type_name or name or label).strip()
        parts = [lead if lead.lower() != label.lower() else label]

        if mat_str:
            first_mat = mat_str.split("|")[0].strip()
            if first_mat and first_mat.lower() not in parts[0].lower():
                parts.append(f"in {first_mat}")

        quals = []
        is_ext = eng.get("Is External")
        if is_ext is True:
            quals.append("external")
        elif is_ext is False:
            quals.append("internal")
        if eng.get("Load Bearing") is True:
            quals.append("load-bearing")
        if quals:
            parts.append(f"({', '.join(quals)})")

        sentence = " ".join(parts).strip()
        if storey_label:
            sentence = f"{sentence} — {storey_label}"
        if desc:
            sentence = f"{sentence}. {desc}" if sentence else desc
        return sentence or label
    except Exception:
        return name or ifc_class


# =============================================================================
# 15. HELPERS
# =============================================================================
def _pick(d: Dict[str, float], *keys: str) -> Optional[float]:
    for k in keys:
        if k in d:
            return d[k]
    return None


def _fmt(v) -> Optional[float]:
    if v is None:
        return None
    try:
        r = round(float(v), 6)
        return r if r > 1e-12 else None
    except (TypeError, ValueError):
        return None


def _bbox(geo: dict, axis: int) -> Optional[float]:
    dims = geo.get("bbox_dims")
    return float(dims[axis]) if dims is not None else None


def _bbox_min_xy(geo: dict) -> Optional[float]:
    dims = geo.get("bbox_dims")
    return float(min(dims[0], dims[1])) if dims is not None else None


def _any_pset(psets: dict, key: str) -> Optional[str]:
    for props in psets.values():
        if isinstance(props, dict) and props.get(key) not in (None, "", "N/A", "NOTDEFINED"):
            return str(props[key])
    return None


# =============================================================================
# 16. ELEMENT → ROWS  (returns a LIST — one row per constituent)
# =============================================================================
def element_to_rows(
    element,
    project_name: str,
    unit_scale: float,
    storey_map: Dict[str, str],
    geom_map: Dict[str, dict],
) -> List[Dict[str, object]]:
    """
    Returns one dict per material constituent.  A single-material element
    returns a list of length 1 (same output as v3.0).  A wall with 3 layers
    returns a list of 3 rows, each with a single resolved material name,
    proportioned volume, mass, and layer-specific thickness.
    """
    guid      = element.GlobalId
    ifc_class = element.is_a()
    name      = getattr(element, "Name",        "") or ""
    desc      = getattr(element, "Description", "") or ""

    psets = get_all_psets(element)          # instance + type merged
    mfr   = psets.get("Pset_ManufacturerTypeInformation", {}) or {}

    # ── Element-level data (shared across all constituent rows) ──────────
    type_name    = get_type_name(element)
    storey_label = storey_map.get(guid) or None
    eng          = get_engineering_props(psets)
    cls_sys, cls_code, cls_name = get_classification(element)

    # ── Quantities (shared) ───────────────────────────────────────────────
    qto = get_quantities(element, unit_scale)
    geo = geom_map.get(guid, {})

    qto_vol  = _pick(qto, "netvolume", "grossvolume", "netsolidvolume",
                     "grosssolidvolume", "volume")
    geom_vol = geo.get("geom_volume")

    occ_used = False
    if qto_vol is None and geom_vol is None:
        occ_vol = occ_volume_fallback(element)
        if occ_vol is not None:
            geom_vol = occ_vol
            occ_used = True

    best_vol = qto_vol if qto_vol is not None else geom_vol

    if occ_used:
        vol_check = "OCC fallback"
    elif qto_vol is not None and geom_vol is not None:
        diff = abs(qto_vol - geom_vol) / max(qto_vol, 1e-9)
        vol_check = "✓ OK" if diff < 0.02 else f"⚠ {diff*100:.1f}%"
    elif geom_vol is not None:
        vol_check = "geom only"
    elif qto_vol is not None:
        vol_check = "Qto only"
    else:
        vol_check = "—"

    area   = _fmt(_pick(qto, "netsidearea", "grosssidearea", "netfloorarea",
                        "grossfloorarea", "area", "netsurfacearea", "netsurface")
                  or geo.get("surface_area"))
    length = _fmt(_pick(qto, "length", "netlength", "grosslength") or _bbox(geo, 0))
    height = _fmt(_pick(qto, "height", "depth")                    or _bbox(geo, 2))
    width  = _fmt(_pick(qto, "breadth", "width", "flangewidth")    or _bbox(geo, 1))
    diam   = _fmt(_pick(qto, "diameter", "outerdiameter", "outercircumference"))

    qto_thickness = _fmt(_pick(qto, "thickness", "width") or _bbox_min_xy(geo))

    # ── Weight from QTO (total element weight) ────────────────────────────
    qto_weight = _pick(qto, "netweight", "grossweight", "weight")

    # ── Lifecycle / circularity psets ─────────────────────────────────────
    lc   = {}
    cond = {}
    for pname in LIFECYCLE_PSETS:
        lc.update(psets.get(pname, {}) or {})
    for pname in ("Pset_Condition", "Pset_PhysicalCondition"):
        cond.update(psets.get(pname, {}) or {})

    det_conn, det_acc, det_int, det_edge = get_detachability(psets)

    # ── Description (uses full assembly for description context) ─────────
    # constituents resolve below, but build description once using mat name
    # from the first constituent; updates per-row are trivial if needed.
    constituents = get_material_constituents(element, unit_scale)
    first_mat_name = (constituents[0]["name"] or "") if constituents else ""
    description = build_description(ifc_class, name, desc, type_name,
                                     first_mat_name, storey_label, eng)

    n_const = len(constituents)
    rows: List[Dict[str, object]] = []

    for idx, const in enumerate(constituents):
        const_density = get_constituent_density(const)
        const_frac    = const["fraction"]

        # Proportioned volume and mass per constituent
        const_vol  = (best_vol * const_frac) if best_vol is not None else None
        if qto_weight is not None:
            const_mass = qto_weight * const_frac
        elif const_vol is not None and const_density is not None:
            const_mass = const_vol * const_density
        else:
            const_mass = None

        # Thickness: per-layer value if available, else fall back to QTO total
        const_thick = const.get("thickness_m") or qto_thickness

        row: Dict[str, object] = {col: None for col in ALL_COLS}

        # ── Identification ────────────────────────────────────────────────
        row["GMAP Id"]        = generate_gmap_id(project_name, guid, idx, n_const)
        row["IFC GUID"]       = guid
        row["GTIN"]           = mfr.get("GlobalTradeItemNumber") or mfr.get("GTIN")
        row["Article Number"] = mfr.get("ArticleNumber") or mfr.get("ModelReference")
        row["External DB Id"] = None

        # ── Element ───────────────────────────────────────────────────────
        row["IFC Class"]      = ifc_class
        row["Type Name"]      = type_name
        row["Element Name"]   = name or None
        row["Description"]    = description
        row["Floor / Storey"] = storey_label
        row["Status"]         = eng.get("Status")

        # ── Material Composition (SINGLE resolved name per row) ───────────
        row["Material / Product"] = const["name"]           # ← single name, no pipes
        row["Material Type"]      = const["mat_type"]       # e.g. IfcMaterialLayer
        row["Layer Build-up"]     = const["full_assembly"] or None  # full assembly context
        row["Material Category"]  = const["category"] or None

        # ── Quantities ────────────────────────────────────────────────────
        row["Volume (m\u00b3)"]      = _fmt(const_vol)
        row["Geom Volume (m\u00b3)"] = _fmt(geom_vol * const_frac if geom_vol is not None else None)
        row["Vol Check"]             = vol_check
        row["Area (m\u00b2)"]        = area       # element-level, same for all constituent rows
        row["Length (m)"]            = length
        row["Height (m)"]            = height
        row["Width (m)"]             = width
        row["Thickness (m)"]         = const_thick
        row["Diameter (m)"]          = diam

        # ── Mass & Carbon ─────────────────────────────────────────────────
        row["Mass (kg)"]             = _fmt(const_mass)
        row["Density (kg/m\u00b3)"]  = _fmt(const_density)
        row["Embodied Carbon A1-A3 (kg CO\u2082e)"] = None
        row["GWP / kg (kg CO\u2082e/kg)"]           = None

        # ── Engineering ───────────────────────────────────────────────────
        row["Is External"]    = eng.get("Is External")
        row["Load Bearing"]   = eng.get("Load Bearing")
        row["Fire Rating"]    = eng.get("Fire Rating")
        row["Thermal Transmittance (W/m\u00b2K)"] = eng.get("Thermal Transmittance (W/m\u00b2K)")
        row["Acoustic Rating (dB)"]               = eng.get("Acoustic Rating (dB)")

        # ── Classification ────────────────────────────────────────────────
        row["Classification System"] = cls_sys  or None
        row["Classification Code"]   = cls_code or None
        row["Classification Name"]   = cls_name or None

        # ── Circularity & EOL ─────────────────────────────────────────────
        row["% Reused"]                   = lc.get("PercentageReused")    or lc.get("Reused")
        row["% Available for Reuse"]      = lc.get("PercentageAvailableForReuse") or lc.get("AvailableForReuse")
        row["Assumed Construction Waste"] = lc.get("AssumedConstructionWaste")
        row["Waste Codes"]                = lc.get("WasteCodes")         or lc.get("WasteCode")
        row["Detachability \u2013 Connection"]    = det_conn
        row["Detachability \u2013 Accessibility"] = det_acc
        row["Detachability \u2013 Intersection"]  = det_int
        row["Detachability \u2013 Product Edge"]  = det_edge

        # ── Lifecycle & Asset ─────────────────────────────────────────────
        row["Lifespan (Years)"]    = (lc.get("ExpectedLifetime") or lc.get("DesignLife")
                                      or lc.get("ServiceLifeDuration")
                                      or lc.get("ExpectedServiceLife")
                                      or lc.get("ServiceLife"))
        row["Installation Date"]   = (mfr.get("InstallationDate")   or mfr.get("ProductionYear")
                                      or mfr.get("ManufactureDate")  or mfr.get("YearOfInstallation"))
        row["Serial Number"]       = (mfr.get("SerialNumber") or mfr.get("BatchReference"))
        row["Technical Condition"] = (cond.get("AssessmentCondition") or cond.get("Condition")
                                      or cond.get("PhysicalCondition"))
        row["Aesthetic Condition"] = (cond.get("AssessmentDescription") or cond.get("AestheticCondition"))
        row["Comment"]             = (cond.get("AssessmentNotes")
                                      or _any_pset(psets, "Comment")
                                      or _any_pset(psets, "Notes")
                                      or _any_pset(psets, "Remarks"))

        rows.append(row)

    return rows


# =============================================================================
# 17. EXCEL WRITER  (format + colours UNCHANGED from v3.0)
# =============================================================================
def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def write_excel(
    rows: List[Dict[str, object]],
    output_path: Path,
    project_name: str,
    schema: str,
    unit_scale: float,
    building_name: str,
    ifc_path: Path,
    dominant_cls_system: str,
    generated_ts: str,
    n_elements: int,
) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Material Passport"
    _passport_sheet(ws1, rows, project_name, schema, building_name, ifc_path,
                    dominant_cls_system, generated_ts)

    ws2 = wb.create_sheet("Summary")
    _summary_sheet(ws2, rows, project_name, schema, unit_scale, building_name,
                   ifc_path, n_elements)

    wb.save(output_path)


def _passport_sheet(ws, rows, project_name, schema, building_name, ifc_path,
                    dominant_cls_system, generated_ts):
    border  = _thin_border()
    n_cols  = len(ALL_COLS)

    col_light: Dict[int, str] = {}
    col_dark:  Dict[int, str] = {}
    idx = 0
    for _grp, (dark, light, cols) in COLUMN_GROUPS.items():
        for _ in cols:
            col_light[idx] = light
            col_dark[idx]  = dark
            idx += 1

    # Row 1 — project banner
    banner_bits = [f"GMAP Material Passport  ·  {project_name}"]
    if building_name:
        banner_bits.append(building_name)
    banner_bits.append(f"IFC Schema: {schema}")
    if dominant_cls_system:
        banner_bits.append(f"Classification: {dominant_cls_system}")
    banner_bits.append(f"File: {ifc_path.name}")
    banner_bits.append(f"Generated: {generated_ts}")
    banner = "  ·  ".join(banner_bits)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=banner)
    c.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Row 2 — group colour bands
    ci = 1
    for grp_name, (dark, light, cols) in COLUMN_GROUPS.items():
        span = len(cols)
        cell = ws.cell(row=2, column=ci, value=grp_name.upper())
        cell.font      = Font(name="Calibri", bold=True, size=8, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=dark)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        if span > 1:
            ws.merge_cells(start_row=2, start_column=ci,
                           end_row=2, end_column=ci + span - 1)
        ci += span
    ws.row_dimensions[2].height = 16

    # Row 3 — column headers
    for ci, col_name in enumerate(ALL_COLS, start=1):
        light = col_light.get(ci - 1, "DDDDDD")
        cell  = ws.cell(row=3, column=ci, value=col_name)
        cell.font      = Font(name="Calibri", bold=True, size=8, color="1A1A1A")
        cell.fill      = PatternFill("solid", fgColor=light)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[3].height = 40

    # Data rows
    for ri, row_data in enumerate(rows):
        er     = ri + 4
        alt_bg = "FFFFFF" if ri % 2 == 0 else "F0F4FB"
        max_lines = 1

        for ci, col_name in enumerate(ALL_COLS, start=1):
            val = row_data.get(col_name)

            if isinstance(val, bool):
                val = "Yes" if val else "No"
            if isinstance(val, str) and len(val) > EXCEL_CELL_CHAR_LIMIT:
                val = val[:EXCEL_CELL_CHAR_LIMIT] + "…[truncated]"

            cell = ws.cell(row=er, column=ci, value=val)
            cell.font   = Font(name="Calibri", size=8)
            cell.border = border

            if col_name == "Vol Check":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val and "✓" in str(val):
                    cell.fill = PatternFill("solid", fgColor="D9EAD3")
                    cell.font = Font(name="Calibri", size=8, color="1D6A54")
                elif val and "⚠" in str(val):
                    cell.fill = PatternFill("solid", fgColor="FCE4D6")
                    cell.font = Font(name="Calibri", size=8, color="78281F")
                else:
                    cell.fill = PatternFill("solid", fgColor=alt_bg)
                continue

            if col_name in WRAP_COLS:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.fill = PatternFill("solid", fgColor=alt_bg)
                text = str(val) if val else ""
                width = COL_WIDTHS.get(col_name, 20)
                chars_per_line = max(int(width * 1.8), 10)
                lines_needed = max(1, -(-len(text) // chars_per_line))
                max_lines = max(max_lines, lines_needed)
                continue

            if col_name in NUMERIC_COLS and isinstance(val, (int, float)):
                cell.alignment    = Alignment(horizontal="right", vertical="center")
                cell.number_format = NUM_FMT.get(col_name, "0.00000")
            elif col_name in {"Is External", "Load Bearing"}:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=False)
            cell.fill = PatternFill("solid", fgColor=alt_bg)

        ws.row_dimensions[er].height = max(15, min(max_lines * 13, 120))

    for ci, col_name in enumerate(ALL_COLS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col_name, 16)

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}3"


def _summary_sheet(ws, rows, project_name, schema, unit_scale, building_name, ifc_path,
                   n_elements: int):

    def _lbl(r, c, key, val, bold_val=False):
        kc = ws.cell(row=r, column=c, value=key)
        kc.font = Font(name="Calibri", bold=True, size=9, color="1F3864")
        vc = ws.cell(row=r, column=c + 1, value=val)
        vc.font = Font(name="Calibri", size=9, bold=bold_val)

    def _h(r, c, val):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "GMAP Material Passport — Summary"
    t.font  = Font(name="Calibri", bold=True, size=14, color="1F3864")
    ws.row_dimensions[1].height = 24

    _lbl(3, 1, "Project:",               project_name)
    _lbl(4, 1, "Building:",              building_name or "—")
    _lbl(5, 1, "IFC Schema:",            schema, bold_val=True)
    _lbl(6, 1, "File:",                  ifc_path.name)
    _lbl(7, 1, "Unit Scale:",            f"{unit_scale} (× QTO values only; geom already SI)")
    _lbl(8, 1, "Total Elements (IFC):",  n_elements, bold_val=True)
    _lbl(9, 1, "Total Passport Rows:",   len(rows),  bold_val=True)

    total_vol  = sum((r.get("Volume (m\u00b3)") or 0)  for r in rows)
    total_mass = sum((r.get("Mass (kg)")         or 0)  for r in rows)
    ok_mat    = sum(1 for r in rows if r.get("Material / Product"))
    ok_vol    = sum(1 for r in rows if r.get("Volume (m\u00b3)") is not None)
    ok_mass   = sum(1 for r in rows if r.get("Mass (kg)") is not None)
    ok_cls    = sum(1 for r in rows if r.get("Classification Code"))
    ok_geo    = sum(1 for r in rows if r.get("Geom Volume (m\u00b3)") is not None)
    ok_storey = sum(1 for r in rows if r.get("Floor / Storey"))
    ok_fire   = sum(1 for r in rows if r.get("Fire Rating"))
    ok_therm  = sum(1 for r in rows if r.get("Thermal Transmittance (W/m\u00b2K)") is not None)

    _lbl(10, 1, "Total Volume (m³):",             round(total_vol,  3))
    _lbl(11, 1, "Total Mass (kg):",               round(total_mass, 1))
    _lbl(12, 1, "Rows with Material:",             f"{ok_mat} / {len(rows)}")
    _lbl(13, 1, "Rows with Volume:",               f"{ok_vol} / {len(rows)}")
    _lbl(14, 1, "Rows with Mass:",                 f"{ok_mass} / {len(rows)}")
    _lbl(15, 1, "Rows with Geometry:",             f"{ok_geo} / {len(rows)}")
    _lbl(16, 1, "Rows with Classification:",       f"{ok_cls} / {len(rows)}")
    _lbl(17, 1, "Rows with Storey assigned:",      f"{ok_storey} / {len(rows)}")
    _lbl(18, 1, "Rows with Fire Rating:",          f"{ok_fire} / {len(rows)}")
    _lbl(19, 1, "Rows with Thermal U-Value:",      f"{ok_therm} / {len(rows)}")

    # Material breakdown
    _h(21, 1, "Material / Product")
    _h(21, 2, "Rows")
    _h(21, 3, "Volume (m³)")
    _h(21, 4, "Mass (kg)")
    _h(21, 5, "Category")

    mat_data: Dict[str, dict] = defaultdict(
        lambda: {"count": 0, "vol": 0.0, "mass": 0.0, "cat": ""}
    )
    for r in rows:
        m = r.get("Material / Product") or "— no material —"
        mat_data[m]["count"] += 1
        mat_data[m]["vol"]   += (r.get("Volume (m\u00b3)") or 0)
        mat_data[m]["mass"]  += (r.get("Mass (kg)")         or 0)
        if not mat_data[m]["cat"]:
            mat_data[m]["cat"] = r.get("Material Category") or ""

    for i, (mat, d) in enumerate(sorted(mat_data.items(), key=lambda x: -x[1]["count"])):
        ws.cell(row=22+i, column=1, value=mat).font   = Font(name="Calibri", size=9)
        ws.cell(row=22+i, column=2, value=d["count"]).font = Font(name="Calibri", size=9)
        ws.cell(row=22+i, column=3, value=round(d["vol"],  4) or None).font = Font(name="Calibri", size=9)
        ws.cell(row=22+i, column=4, value=round(d["mass"], 2) or None).font = Font(name="Calibri", size=9)
        ws.cell(row=22+i, column=5, value=d["cat"] or None).font = Font(name="Calibri", size=9)

    # IFC class breakdown
    r0 = 24 + len(mat_data)
    _h(r0, 1, "IFC Class"); _h(r0, 2, "Count")
    cls_count = Counter(r.get("IFC Class") or "?" for r in rows)
    for i, (cls, cnt) in enumerate(sorted(cls_count.items(), key=lambda x: -x[1])):
        ws.cell(row=r0+1+i, column=1, value=cls).font  = Font(name="Calibri", size=9)
        ws.cell(row=r0+1+i, column=2, value=cnt).font  = Font(name="Calibri", size=9)

    # Floor / Storey breakdown
    r1 = r0 + 2 + len(cls_count)
    _h(r1, 1, "Floor / Storey"); _h(r1, 2, "Count")
    storey_count = Counter(r.get("Floor / Storey") or "(unassigned)" for r in rows)
    for i, (st, cnt) in enumerate(sorted(storey_count.items(), key=lambda x: -x[1])):
        ws.cell(row=r1+1+i, column=1, value=st).font  = Font(name="Calibri", size=9)
        ws.cell(row=r1+1+i, column=2, value=cnt).font  = Font(name="Calibri", size=9)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18


# =============================================================================
# 18. CSV WRITER
# =============================================================================
def write_csv(rows: List[Dict[str, object]], csv_path: Path) -> None:
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=ALL_COLS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            clean = {k: ("" if v is None else v) for k, v in row.items()}
            writer.writerow(clean)


# =============================================================================
# 19. ROBUST FILE OUTPUT
# =============================================================================
def safe_write(write_fn, target_path: Path, *args, **kwargs) -> Path:
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    stem, suffix, parent = target_path.stem, target_path.suffix, target_path.parent

    attempts = [target_path, parent / f"{stem}_{ts}{suffix}"]
    try:
        attempts.append(Path.home() / "Documents" / f"{stem}_{ts}{suffix}")
    except Exception:
        pass
    try:
        attempts.append(Path(__file__).resolve().parent / f"{stem}_{ts}{suffix}")
    except Exception:
        pass
    attempts.append(Path.cwd() / f"{stem}_{ts}{suffix}")

    last_err: Optional[Exception] = None
    for i, p in enumerate(attempts):
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            write_fn(p, *args, **kwargs)
            if i > 0:
                print(f"  [WARNING] '{target_path.name}' was locked - saved instead as: {p}")
            return p
        except (PermissionError, Exception) as e:
            last_err = e
            continue

    raise RuntimeError(f"Could not write '{target_path.name}' to any location. "
                       f"Last error: {last_err}")


# =============================================================================
# 20. FILE PICKER
# =============================================================================
def pick_ifc_file() -> Path:
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title="Select IFC File",
            filetypes=[("IFC files", "*.ifc *.ifczip *.ifcxml"), ("All files", "*.*")],
        )
        root.destroy()
        if not path:
            print("No file selected.")
            sys.exit(0)
        return Path(path)
    except Exception as e:
        print(f"File picker unavailable ({e}).\nUsage:  python ifc_to_passport_v3_1.py project.ifc")
        sys.exit(1)


# =============================================================================
# 21. MAIN
# =============================================================================
def main() -> None:
    # ── CLI: python ifc_to_excel.py <input.ifc> [output.xlsx] [--no-geom] ──────
    # Flags (position-independent, always prefixed with --):
    #   --no-geom   skip geometry extraction entirely (fast, recommended for server use)
    flags   = {a for a in sys.argv[1:] if a.startswith("--")}
    posargs = [a for a in sys.argv[1:] if not a.startswith("--")]
    no_geom = "--no-geom" in flags

    if not posargs:
        print("Usage: python ifc_to_excel.py <input.ifc> [output.xlsx] [--no-geom]",
              file=sys.stderr)
        print("  <input.ifc>    path to the IFC file (required)", file=sys.stderr)
        print("  [output.xlsx]  destination Excel path (optional; auto-named if omitted)",
              file=sys.stderr)
        print("  --no-geom      skip geometry extraction (recommended for server use)",
              file=sys.stderr)
        sys.exit(1)

    ifc_path = Path(posargs[0])

    if not ifc_path.exists():
        print(f"[ERROR] File not found: {ifc_path}", file=sys.stderr)
        sys.exit(1)

    if not ifc_path.is_file():
        print(f"[ERROR] Not a regular file: {ifc_path}", file=sys.stderr)
        sys.exit(1)

    # When an explicit output path is supplied (server mode), write there directly
    # and skip CSV generation.  When omitted (interactive mode), auto-name both.
    explicit_output = len(posargs) >= 2
    if explicit_output:
        output_xlsx_path = Path(posargs[1])
        output_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    else:
        safe_name_default = "".join(c if c.isalnum() or c in " _-" else "_"
                                    for c in str(ifc_path.stem)).strip() or "Project"
        output_xlsx_path = ifc_path.parent / f"MaterialPassport_{safe_name_default}.xlsx"

    sep = "=" * 66
    print(f"\n{sep}")
    print("GMAP IFC to Material Passport Converter v3.1")
    print(f"{sep}")
    print(f"  Input IFC  : {ifc_path.resolve()}")
    print(f"  Output XLSX: {output_xlsx_path.resolve()}")
    print(f"  Working dir: {Path.cwd()}")
    print(f"  Python     : {sys.version.split()[0]}")
    print(f"  Geom mode  : {'disabled (--no-geom)' if no_geom else 'enabled'}")
    print(f"  File size  : {ifc_path.stat().st_size / 1024:.1f} KB")
    sys.stdout.flush()
    if not _OCC_AVAILABLE:
        print("  [INFO] pythonocc-core not found - OCC volume fallback disabled "
              "(optional).  pip install pythonocc-core")
        sys.stdout.flush()

    try:
        model = ifcopenshell.open(str(ifc_path))
    except Exception as e:
        print(f"[ERROR] Cannot open IFC file: {e}", file=sys.stderr)
        sys.exit(1)

    schema     = getattr(model, "schema", "Unknown")
    unit_scale = get_unit_scale(model)
    print(f"  Schema : {schema}")
    print(f"  Unit scale -> SI : {unit_scale}  (QTO values only; geometry already SI)")

    try:
        projects = model.by_type("IfcProject")
    except Exception:
        projects = []
    project_name  = (getattr(projects[0], "Name", None) if projects else None) or ifc_path.stem
    building_name = get_building_name(model)
    print(f"  Project  : {project_name}")
    print(f"  Building : {building_name or '-'}")

    try:
        counts = Counter(e.is_a() for e in model)
        print(f"\n  Entity inventory (top 20):")
        for k, v in sorted(counts.items(), key=lambda x: -x[1])[:20]:
            print(f"    {k:<42}  {v}")
    except Exception as e:
        print(f"  [WARNING] Entity inventory failed: {e}")

    print(f"\n  Building spatial tree...")
    storey_map = build_storey_map(model)
    print(f"  Tagged {len(storey_map)} elements with storey names")

    if no_geom:
        print("\n  Geometry extraction: skipped (--no-geom flag)")
    else:
        print(f"\n  Geometry extraction: {os.cpu_count() or 4} thread(s) available")
    sys.stdout.flush()
    geom_map = extract_geometry(model, skip_geom=no_geom)

    print(f"\n  Collecting elements...")
    seen: set = set()
    all_elements = []
    for ifc_type in TARGET_IFC_TYPES:
        try:
            for el in model.by_type(ifc_type):
                if el.GlobalId not in seen:
                    seen.add(el.GlobalId)
                    all_elements.append(el)
        except Exception as e:
            print(f"  [WARNING] by_type({ifc_type}): {e}")

    if not all_elements:
        print("  [INFO] No typed elements found - falling back to all IfcElement")
        for el in model.by_type("IfcElement"):
            if el.GlobalId not in seen:
                seen.add(el.GlobalId)
                all_elements.append(el)

    n_elements = len(all_elements)
    print(f"  Found {n_elements} IFC elements")

    # Process — each element may produce multiple constituent rows
    rows: List[Dict[str, object]] = []
    elem_stats: Dict[str, int]    = defaultdict(int)
    total = n_elements

    # In server mode (explicit_output) use newlines so each line flushes through
    # the pipe.  In interactive mode use \r for a compact progress bar.
    log_interval = max(1, min(50, total // 20)) if total else 1
    for i, el in enumerate(all_elements):
        if i % log_interval == 0 or i == total - 1:
            pct = (i + 1) / total * 100 if total else 100
            if explicit_output:
                print(f"  Element {i + 1}/{total} ({pct:.0f}%)...")
            else:
                print(f"  Processing {i+1:>5} / {total}  ({pct:5.1f}%)...", end="\r")
        try:
            el_rows = element_to_rows(el, project_name, unit_scale, storey_map, geom_map)
            rows.extend(el_rows)
            elem_stats[el.is_a()] += 1
        except Exception as ex:
            if not explicit_output:
                print()  # end the \r line before the warning
            print(f"  [WARNING] element {getattr(el, 'GlobalId', '?')} "
                  f"({el.is_a()}): {ex}")
            print(traceback.format_exc(), file=sys.stderr)

    if not explicit_output:
        print()  # newline after interactive progress bar

    vol_total  = sum((r.get("Volume (m\u00b3)") or 0)  for r in rows)
    mass_total = sum((r.get("Mass (kg)")         or 0)  for r in rows)
    mat_ok     = sum(1 for r in rows if r.get("Material / Product"))
    geo_ok     = sum(1 for r in rows if r.get("Geom Volume (m\u00b3)") is not None)
    mass_ok    = sum(1 for r in rows if r.get("Mass (kg)") is not None)
    fire_ok    = sum(1 for r in rows if r.get("Fire Rating"))
    therm_ok   = sum(1 for r in rows if r.get("Thermal Transmittance (W/m\u00b2K)") is not None)

    print(f"\n  IFC elements processed : {n_elements}")
    print(f"  Total passport rows    : {len(rows)}  (one per material constituent)")
    for t, cnt in sorted(elem_stats.items()):
        print(f"    {t:<44}  {cnt}")
    print(f"\n  Total volume (m3) : {vol_total:.4f}")
    print(f"  Total mass   (kg) : {mass_total:.2f}")
    print(f"  With material     : {mat_ok} / {len(rows)}")
    print(f"  With geometry     : {geo_ok} / {len(rows)}")
    print(f"  With mass         : {mass_ok} / {len(rows)}")
    print(f"  With fire rating  : {fire_ok} / {len(rows)}")
    print(f"  With thermal U    : {therm_ok} / {len(rows)}")

    cls_counter = Counter(r.get("Classification System") for r in rows
                          if r.get("Classification System"))
    dominant_cls_system = cls_counter.most_common(1)[0][0] if cls_counter else ""
    generated_ts = datetime.datetime.now().strftime("%d-%b-%Y %H:%M")

    print(f"\n  Writing Excel...")
    print(f"  Target path: {output_xlsx_path.resolve()}")
    sys.stdout.flush()
    try:
        xlsx_path = safe_write(
            lambda p: write_excel(rows, p, project_name, schema, unit_scale,
                                   building_name, ifc_path, dominant_cls_system,
                                   generated_ts, n_elements),
            output_xlsx_path,
        )
    except Exception as save_err:
        print(f"[ERROR] Excel write failed: {save_err}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        sys.exit(1)

    if not xlsx_path.exists():
        print(f"[ERROR] Excel file missing after save: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    xlsx_bytes = xlsx_path.stat().st_size
    print(f"  Excel saved: {xlsx_path.resolve()} ({xlsx_bytes:,} bytes)")
    sys.stdout.flush()

    if not explicit_output:
        safe_name = "".join(c if c.isalnum() or c in " _-" else "_"
                            for c in str(project_name)).strip() or "Project"
        csv_base = ifc_path.parent / f"MaterialPassport_{safe_name}"
        print(f"  Writing CSV...")
        csv_path = safe_write(lambda p: write_csv(rows, p), csv_base.with_suffix(".csv"))
        print(f"  >> {csv_path}")

    print(f"\n{sep}")
    print(f"  [DONE] {n_elements} IFC elements -> {len(rows)} passport rows "
          f"({len(ALL_COLS)} columns)")
    print(f"  >> {xlsx_path}")
    print(f"{sep}\n")
    # Machine-readable marker parsed by the TypeScript caller (ifcProcessingService.ts)
    print(f"XLSX_OUTPUT:{xlsx_path}", flush=True)
    sys.exit(0)


if __name__ == "__main__":
    main()